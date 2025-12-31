
'''
Author: Matthew Bandura
Created on Wednesday 19th November, 16:31:31  
'''

import pandas as pd
import openpyxl

def generate_BSF_reg_status(BSF_cut, last_month_BSF_cut):
    #file is too complex to load normally so must be read_only
    BSF_wb = openpyxl.load_workbook(BSF_cut, read_only= True, data_only= True)
    BSF_tables_DR = BSF_wb['Essential Tables for DR']

    #iterates through the rows containing the registration status table
    BSF_reg_status_data = BSF_tables_DR.iter_rows(min_row = 20, max_row = 29,
                                   min_col = 2, max_col = 3,
                                   values_only=True)
    
    #turns the data into a list rather than a generator type object
    BSF_reg_status_data = list(BSF_reg_status_data)

    #turns the data into a pandas data frame with rows starting at 1 columns 0
    BSF_reg_status_df = pd.DataFrame(BSF_reg_status_data[1:], columns  = BSF_reg_status_data[0])
    BSF_reg_status_df.rename(columns={BSF_reg_status_df.columns[-1]: 'Current Month'}, inplace=True)
    BSF_wb.close()




    last_month_BSF_wb = openpyxl.load_workbook(last_month_BSF_cut, read_only= True, data_only= True)
    last_month_BSF_tables_DR = last_month_BSF_wb['Essential Tables for DR']

    #iterates through the rows containing the registration status table
    last_month_BSF_reg_status_data = last_month_BSF_tables_DR.iter_rows(min_row = 20, max_row = 29,
                                   min_col = 2, max_col = 3,
                                   values_only=True)
    
    #turns the data into a list rather than a generator type object
    last_month_BSF_reg_status_data = list(last_month_BSF_reg_status_data)

    #turns the data into a pandas data frame with rows starting at 1 columns 0
    last_month_BSF_reg_status_df = pd.DataFrame(last_month_BSF_reg_status_data[1:], columns  = last_month_BSF_reg_status_data[0])
    last_month_BSF_reg_status_df.rename(columns={last_month_BSF_reg_status_df.columns[-1]: 'Last Month'}, inplace=True)
    last_month_BSF_wb.close()

    #combine this and last month's into one df
    BSF_reg_status_df_combined = pd.merge(last_month_BSF_reg_status_df, BSF_reg_status_df)
    return BSF_reg_status_df_combined